#!/usr/bin/env python3
"""
Build inat/elevation-bands.json from cp_elevation_bands.xlsx.

These are curated, published elevation ranges (min/max metres) for carnivorous
plant species whose iNaturalist records are largely *obscured*. The climate
pipeline (scripts/fetch-inat-data.js) uses them to clamp the elevation of a
species' obscured cells into its real habitat band before the lapse-rate
temperature correction — sharpening the climate envelope for species whose
obscured coordinates would otherwise land at the wrong elevation.

To add/adjust data: edit the spreadsheet and re-run this script. Columns are read
by header name, so column order can change. To support a new genus, just add a
sheet (or rows) with the same headers — no code change needed.

    python scripts/build-elevation-bands.py
"""
import json
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "cp_elevation_bands.xlsx"
OUT = ROOT / "inat" / "elevation-bands.json"

# Sheets to read (genus data). The README sheet is skipped automatically.
SHEETS = ["Heliamphora", "Pinguicula", "Utricularia", "Drosera (montane)"]

# A real binomial: "Genus epithet" (epithet lowercase, may be hyphenated).
# This skips subgenus/section placeholder rows and "Genus sp. 'Tepui'" entries.
BINOMIAL = re.compile(r"^[A-Z][a-z]+ [a-z]+(?:-[a-z]+)?$")


def to_num(v):
    if v is None:
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def main():
    if not XLSX.exists():
        sys.exit(f"Source spreadsheet not found: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    species = {}
    skipped, dupes = [], []
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            print(f"  ! sheet not found, skipping: {sheet}")
            continue
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        try:
            i_name = header.index("Species")
            i_min = header.index("Min elevation (m)")
            i_max = header.index("Max elevation (m)")
        except ValueError as e:
            sys.exit(f"Sheet '{sheet}' is missing an expected column: {e}")

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[i_name]
            if not name:
                continue
            name = str(name).strip()
            mn, mx = to_num(row[i_min]), to_num(row[i_max])
            if not BINOMIAL.match(name):
                continue  # subgenus/section placeholder or "sp." form
            if mn is None and mx is None:
                skipped.append(name)  # no numeric bound yet
                continue
            if mn is not None and mx is not None and mn > mx:
                mn, mx = mx, mn  # tolerate reversed entries
            key = name.lower()
            if key in species:
                dupes.append(name)
            species[key] = [mn, mx]

    out = {
        "_about": (
            "Curated published elevation bands [minMetres, maxMetres] (either may be "
            "null) for carnivorous plant species. Used by scripts/fetch-inat-data.js to "
            "clamp obscured-observation elevations into the species' real habitat band "
            "before the lapse-rate temperature correction. Keys are lowercase 'genus "
            "species'. Regenerate with scripts/build-elevation-bands.py."
        ),
        "_generated": date.today().isoformat(),
        "_count": len(species),
        "species": dict(sorted(species.items())),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, indent=0, ensure_ascii=True), encoding="utf-8")
    print(f"Wrote {OUT}: {len(species)} species with bands.")
    if skipped:
        print(f"  ({len(skipped)} binomial rows skipped — no numeric bound yet)")
    if dupes:
        print(f"  ! duplicate species across sheets (last wins): {sorted(set(dupes))}")


if __name__ == "__main__":
    main()
